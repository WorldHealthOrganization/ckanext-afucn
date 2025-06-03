import ckan.plugins.toolkit as tk
from ckan.common import config
import os
import logging
from pathlib import Path
from openpyxl import load_workbook
import shutil
import tempfile


log = logging.getLogger(__name__)

try:
    storage_path = config.get('ckan.storage_path')
except:
    log.critical('''Please specify a ckan.storage_path in your config
                         for your uploads''')


def get_resource_path(resource_id):

    resource_path = storage_path + "/resources/" + resource_id[0:3] + "/" + resource_id[3:6] + "/" + resource_id[6:]
    
    return resource_path

def create_subresource(context, resource_dict):

    package_data = {
        'id': resource_dict['package_id'],
    }
    
    resources_list = tk.get_action("package_show")(context, package_data)['resources']
    
    subresource_name = resource_dict['name'].rsplit('.', 1)[0] + '.geojson'
    names_list = []

    for i in resources_list:
        names_list.append(i['name'])
    if subresource_name not in names_list:

        data_dict = {
            'package_id': resource_dict['package_id'],
            'name': subresource_name,
            'url': subresource_name,
            'url_type': 'upload',
        }

        resource_path = get_resource_path(resource_dict['id'])
        resource_copy_path = resource_path + ".xlsx"
        
        # Copy the file to the same location with .xls extension
        shutil.copy(resource_path, resource_copy_path)

        # Load the excel file from copied resource
        wb = load_workbook(resource_copy_path)
        ws = wb['Sheet1']

        for idx, row in enumerate(ws.iter_rows()):
            geojson_text = row[1].value

        f = tempfile.NamedTemporaryFile(mode='w+t', delete=False)
        f.write(geojson_text)
        file_name = f.name
        f.close()
        
        # Create the map resource from the excel file and
        x = tk.get_action("resource_create")(context, data_dict)
        upload_path = storage_path + '/resources/' + x['id'][0:3] + "/" + x['id'][3:6]
        upload_filename = x['id'][6:]
        Path(upload_path).mkdir(parents=True, exist_ok=True)
        filepath = Path(os.path.join(upload_path, upload_filename))
        shutil.copy(file_name, filepath)
        
        # remove temp xls resource from resources directory 
        os.remove(resource_copy_path)
        # remove temp subresource file
        os.remove(file_name) 
    
        return resource_dict

    else:
        print("======================================================================")
        print('subresource not created')
        print("======================================================================")
        
        return resource_dict

